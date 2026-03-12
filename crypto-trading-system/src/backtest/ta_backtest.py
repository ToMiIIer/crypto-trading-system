"""Deterministic TA-only backtest workflow and report export."""

from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Sequence

from openpyxl import Workbook

from src.backtest.history_downloader import parse_date_to_utc_ms
from src.ta.deterministic_ta import combine_vote_score, compute_indicators, compute_signals, derive_action

TRADE_HEADERS = [
    "trade_id",
    "symbol",
    "interval",
    "entry_time",
    "exit_time",
    "entry_price",
    "exit_price",
    "quantity",
    "entry_notional",
    "entry_fee",
    "exit_fee",
    "pnl_abs",
    "pnl_pct",
    "exit_reason",
    "bars_held",
]

EQUITY_HEADERS = [
    "open_time",
    "close_time",
    "open",
    "high",
    "low",
    "close",
    "signal_action",
    "signal_score",
    "signal_confidence",
    "cash",
    "equity",
    "position_qty",
]


@dataclass(slots=True)
class OpenPosition:
    entry_time: int
    entry_price: float
    quantity: float
    entry_notional: float
    entry_fee: float
    stop_loss_price: float
    take_profit_price: float
    entry_index: int


def load_candles_csv(path: Path | str) -> list[dict[str, Any]]:
    csv_path = Path(path)
    with csv_path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        rows: list[dict[str, Any]] = []
        for row in reader:
            rows.append(
                {
                    "open_time": int(row["open_time"]),
                    "open": float(row["open"]),
                    "high": float(row["high"]),
                    "low": float(row["low"]),
                    "close": float(row["close"]),
                    "volume": float(row["volume"]),
                    "close_time": int(row["close_time"]),
                }
            )
    rows.sort(key=lambda item: int(item["open_time"]))
    return rows


def slice_candles(
    candles: Sequence[dict[str, Any]],
    *,
    start: str | None = None,
    end: str | None = None,
    years: int = 6,
) -> list[dict[str, Any]]:
    ordered = sorted((dict(row) for row in candles), key=lambda item: int(item["open_time"]))
    if not ordered:
        return []

    start_ms = parse_date_to_utc_ms(start) if start else None
    end_ms = parse_date_to_utc_ms(end, end_of_day=True) if end else None
    if start_ms is None and end_ms is None:
        end_ms = int(ordered[-1]["open_time"]) + 1
        start_ms = end_ms - (max(years, 1) * 365 * 24 * 60 * 60 * 1000)
    elif start_ms is None:
        start_ms = ordered[0]["open_time"]
    elif end_ms is None:
        end_ms = int(ordered[-1]["open_time"]) + 1

    if start_ms >= end_ms:
        raise ValueError("start must be earlier than end")

    return [row for row in ordered if start_ms <= int(row["open_time"]) < end_ms]


def _timestamp_to_iso(timestamp_ms: int) -> str:
    return datetime.fromtimestamp(timestamp_ms / 1000, tz=timezone.utc).isoformat()


def _max_drawdown_pct(equity_rows: Sequence[dict[str, Any]]) -> float:
    peak = 0.0
    max_drawdown = 0.0
    for row in equity_rows:
        equity = float(row["equity"])
        peak = max(peak, equity)
        if peak <= 0:
            continue
        drawdown = ((peak - equity) / peak) * 100.0
        max_drawdown = max(max_drawdown, drawdown)
    return round(max_drawdown, 6)


def _close_position(
    *,
    position: OpenPosition,
    exit_price: float,
    exit_time: int,
    exit_reason: str,
    current_index: int,
    fee_rate: float,
) -> tuple[dict[str, Any], float]:
    gross_exit_notional = position.quantity * exit_price
    exit_fee = gross_exit_notional * fee_rate
    pnl_abs = gross_exit_notional - exit_fee - position.entry_notional - position.entry_fee
    pnl_pct = (pnl_abs / position.entry_notional * 100.0) if position.entry_notional > 0 else 0.0
    trade = {
        "trade_id": f"{position.entry_time}-{exit_time}",
        "entry_time": _timestamp_to_iso(position.entry_time),
        "exit_time": _timestamp_to_iso(exit_time),
        "entry_price": round(position.entry_price, 8),
        "exit_price": round(exit_price, 8),
        "quantity": round(position.quantity, 8),
        "entry_notional": round(position.entry_notional, 8),
        "entry_fee": round(position.entry_fee, 8),
        "exit_fee": round(exit_fee, 8),
        "pnl_abs": round(pnl_abs, 8),
        "pnl_pct": round(pnl_pct, 8),
        "exit_reason": exit_reason,
        "bars_held": max(1, current_index - position.entry_index + 1),
    }
    cash_delta = gross_exit_notional - exit_fee
    return trade, cash_delta


def backtest_ta(
    *,
    candles: Sequence[dict[str, Any]],
    ta_config: dict[str, Any],
    symbol: str,
    interval: str,
    start: str | None = None,
    end: str | None = None,
    years: int = 6,
    initial_capital: float = 10000.0,
    size_pct: float = 1.0,
    sl_pct: float = 0.02,
    tp_pct: float = 0.04,
    fee_bps: float = 4.0,
    sl_tp_priority: str = "stop_first",
) -> dict[str, Any]:
    if sl_tp_priority not in {"stop_first", "tp_first"}:
        raise ValueError("sl_tp_priority must be one of: stop_first, tp_first")

    selected = slice_candles(candles, start=start, end=end, years=years)
    if len(selected) < 2:
        raise ValueError("backtest requires at least 2 candles in the selected period")

    fee_rate = fee_bps / 10000.0
    cash = float(initial_capital)
    position: OpenPosition | None = None
    pending_action: str | None = None
    trades: list[dict[str, Any]] = []
    equity_rows: list[dict[str, Any]] = []

    for index, candle in enumerate(selected):
        current_open = float(candle["open"])
        current_high = float(candle["high"])
        current_low = float(candle["low"])
        current_close = float(candle["close"])
        current_open_time = int(candle["open_time"])
        current_close_time = int(candle["close_time"])

        if pending_action == "BUY" and position is None:
            entry_notional = (cash * size_pct) / (1.0 + fee_rate)
            if entry_notional > 0:
                quantity = entry_notional / max(current_open, 1e-9)
                entry_fee = entry_notional * fee_rate
                cash -= entry_notional + entry_fee
                position = OpenPosition(
                    entry_time=current_open_time,
                    entry_price=current_open,
                    quantity=quantity,
                    entry_notional=entry_notional,
                    entry_fee=entry_fee,
                    stop_loss_price=current_open * (1.0 - sl_pct),
                    take_profit_price=current_open * (1.0 + tp_pct),
                    entry_index=index,
                )
        elif pending_action == "SELL" and position is not None:
            trade, cash_delta = _close_position(
                position=position,
                exit_price=current_open,
                exit_time=current_open_time,
                exit_reason="signal_sell",
                current_index=index,
                fee_rate=fee_rate,
            )
            trades.append(
                {
                    "symbol": symbol.upper(),
                    "interval": interval,
                    **trade,
                }
            )
            cash += cash_delta
            position = None

        if position is not None:
            hit_stop = current_low <= position.stop_loss_price
            hit_take = current_high >= position.take_profit_price
            if hit_stop or hit_take:
                if hit_stop and hit_take:
                    exit_reason = "stop_loss" if sl_tp_priority == "stop_first" else "take_profit"
                    exit_price = position.stop_loss_price if sl_tp_priority == "stop_first" else position.take_profit_price
                elif hit_stop:
                    exit_reason = "stop_loss"
                    exit_price = position.stop_loss_price
                else:
                    exit_reason = "take_profit"
                    exit_price = position.take_profit_price

                trade, cash_delta = _close_position(
                    position=position,
                    exit_price=exit_price,
                    exit_time=current_close_time,
                    exit_reason=exit_reason,
                    current_index=index,
                    fee_rate=fee_rate,
                )
                trades.append(
                    {
                        "symbol": symbol.upper(),
                        "interval": interval,
                        **trade,
                    }
                )
                cash += cash_delta
                position = None

        window = selected[: index + 1]
        indicators = compute_indicators(window, ta_config)
        signals = compute_signals(indicators, window, ta_config)
        vote_score, confidence, _breakdown = combine_vote_score(signals, ta_config)
        pending_action = derive_action(vote_score, ta_config)

        equity = cash if position is None else cash + (position.quantity * current_close)
        equity_rows.append(
            {
                "open_time": _timestamp_to_iso(current_open_time),
                "close_time": _timestamp_to_iso(current_close_time),
                "open": round(current_open, 8),
                "high": round(current_high, 8),
                "low": round(current_low, 8),
                "close": round(current_close, 8),
                "signal_action": pending_action,
                "signal_score": round(vote_score, 8),
                "signal_confidence": round(confidence, 8),
                "cash": round(cash, 8),
                "equity": round(equity, 8),
                "position_qty": round(position.quantity, 8) if position is not None else 0.0,
            }
        )

    wins = sum(1 for trade in trades if float(trade["pnl_abs"]) > 0)
    losses = sum(1 for trade in trades if float(trade["pnl_abs"]) < 0)
    total_trades = len(trades)
    total_pnl_abs = sum(float(trade["pnl_abs"]) for trade in trades)
    final_equity = float(equity_rows[-1]["equity"]) if equity_rows else float(initial_capital)
    summary = {
        "symbol": symbol.upper(),
        "interval": interval,
        "start": equity_rows[0]["open_time"] if equity_rows else "",
        "end": equity_rows[-1]["close_time"] if equity_rows else "",
        "initial_capital": round(float(initial_capital), 8),
        "final_equity": round(final_equity, 8),
        "total_pnl_abs": round(total_pnl_abs, 8),
        "total_return_pct": round(((final_equity - initial_capital) / initial_capital * 100.0), 8),
        "total_trades": total_trades,
        "wins": wins,
        "losses": losses,
        "win_rate": round((wins / total_trades) if total_trades else 0.0, 8),
        "max_drawdown_pct": _max_drawdown_pct(equity_rows),
        "open_position": position is not None,
        "size_pct": float(size_pct),
        "sl_pct": float(sl_pct),
        "tp_pct": float(tp_pct),
        "fee_bps": float(fee_bps),
        "sl_tp_priority": sl_tp_priority,
    }

    return {
        "trades": trades,
        "equity": equity_rows,
        "summary": summary,
    }


def _write_csv(path: Path, headers: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({header: row.get(header, "") for header in headers})


def _write_summary_json(path: Path, summary: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(summary, indent=2), encoding="utf-8")


def _write_excel(path: Path, *, trades: list[dict[str, Any]], equity: list[dict[str, Any]], summary: dict[str, Any]) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["metric", "value"])
    for key, value in summary.items():
        summary_sheet.append([key, value])

    trades_sheet = workbook.create_sheet("Trades")
    trades_sheet.append(TRADE_HEADERS)
    for row in trades:
        trades_sheet.append([row.get(header, "") for header in TRADE_HEADERS])

    equity_sheet = workbook.create_sheet("Equity")
    equity_sheet.append(EQUITY_HEADERS)
    for row in equity:
        equity_sheet.append([row.get(header, "") for header in EQUITY_HEADERS])

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def export_backtest_artifacts(result: dict[str, Any], reports_dir: Path | str) -> dict[str, str]:
    output_dir = Path(reports_dir).resolve()
    trades_csv = output_dir / "trades.csv"
    equity_csv = output_dir / "equity.csv"
    summary_json = output_dir / "summary.json"
    xlsx_path = output_dir / "backtest.xlsx"

    _write_csv(trades_csv, TRADE_HEADERS, list(result.get("trades", [])))
    _write_csv(equity_csv, EQUITY_HEADERS, list(result.get("equity", [])))
    _write_summary_json(summary_json, dict(result.get("summary", {})))
    _write_excel(
        xlsx_path,
        trades=list(result.get("trades", [])),
        equity=list(result.get("equity", [])),
        summary=dict(result.get("summary", {})),
    )

    return {
        "trades_csv": str(trades_csv),
        "equity_csv": str(equity_csv),
        "summary_json": str(summary_json),
        "xlsx": str(xlsx_path),
    }
