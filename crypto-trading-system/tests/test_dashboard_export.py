from __future__ import annotations

from datetime import datetime, timedelta, timezone
from pathlib import Path
import tempfile
import unittest

from openpyxl import load_workbook

from src.storage.dashboard_exporter import DashboardExporter
from src.storage.repository import StorageRepository


class DashboardPersistenceTests(unittest.TestCase):
    def setUp(self) -> None:
        self._tmpdir = tempfile.TemporaryDirectory()
        self.tmp_path = Path(self._tmpdir.name)
        self.db_path = self.tmp_path / "dashboard_test.db"
        self.repo = StorageRepository(f"sqlite:///{self.db_path}")
        self.repo.initialize()

    def tearDown(self) -> None:
        self._tmpdir.cleanup()

    @staticmethod
    def _sample_payload(run_id: str) -> dict[str, object]:
        return {
            "run_id": run_id,
            "pair": "BTC/USDC",
            "timeframe": "4h",
            "status": "NO_TRADE",
            "indicators": {
                "rsi_14": 53.2,
                "ema_21": 67000.0,
                "ema_50": 66800.0,
                "ema_200": 64000.0,
                "atr_14": 500.0,
                "macd": {"value": 12.0, "signal": 10.0, "histogram": 2.0},
                "bollinger_20_2": {"middle": 66500.0, "upper": 68000.0, "lower": 65000.0},
            },
            "hypotheses": [
                {
                    "run_id": run_id,
                    "pair": "BTC/USDC",
                    "timeframe": "4h",
                    "agent_id": "llm_analyst",
                    "action": "HOLD",
                    "confidence": 0.5,
                    "reasoning": "llm_disabled_missing_api_key:openai",
                    "risk_notes": "LLM disabled fallback.",
                    "provider_used": "disabled",
                    "error_code": "llm_disabled_missing_api_key",
                    "error_message": "missing credentials for provider 'openai'",
                }
            ],
            "consensus": {
                "action": "HOLD",
                "weighted_confidence": 0.5,
                "threshold_passed": False,
                "reasoning": "not actionable",
            },
            "risk_decision": {
                "approved": False,
                "action": "HOLD",
                "reason": "consensus_not_actionable",
                "position_pct": 0.0,
                "stop_loss_pct": 0.0,
            },
            "execution": {"status": "NO_TRADE", "reason": "consensus_not_actionable"},
            "warnings": ["llm_fallback"],
            "errors": [],
        }

    def test_persist_dashboard_run_writes_pipeline_and_agent_rows(self) -> None:
        run_id = "run-persist-1"
        started = datetime.now(timezone.utc)
        finished = started + timedelta(seconds=2)
        payload = self._sample_payload(run_id)

        self.repo.persist_dashboard_run(payload=payload, started_at=started, finished_at=finished)

        runs = self.repo.list_pipeline_runs()
        self.assertEqual(len(runs), 1)
        self.assertEqual(runs[0]["run_id"], run_id)
        self.assertEqual(runs[0]["status"], "NO_TRADE")
        self.assertGreaterEqual(int(runs[0]["warnings_count"]), 1)

        agents = self.repo.list_agent_outputs()
        self.assertEqual(len(agents), 1)
        self.assertEqual(agents[0]["agent_id"], "llm_analyst")
        self.assertEqual(agents[0]["provider_used"], "disabled")
        self.assertEqual(agents[0]["error_code"], "llm_disabled_missing_api_key")

        indicators = self.repo.list_indicators()
        self.assertEqual(len(indicators), 1)
        self.assertAlmostEqual(float(indicators[0]["rsi_14"]), 53.2, places=4)

    def test_export_dashboard_creates_xlsx_and_csv_outputs(self) -> None:
        run_id = "run-export-1"
        now = datetime.now(timezone.utc)
        self.repo.persist_dashboard_run(
            payload=self._sample_payload(run_id),
            started_at=now,
            finished_at=now + timedelta(seconds=1),
        )
        position = self.repo.open_position(
            run_id=run_id,
            pair="BTC/USDC",
            timeframe="4h",
            side="LONG",
            entry_price=68000.0,
            size_pct=0.10,
            quantity=0.01,
            stop_loss_price=67000.0,
            take_profit_price=69000.0,
            reason="paper_open",
        )
        self.repo.refresh_performance(run_id=run_id, pair="BTC/USDC", timeframe="4h")
        self.repo.close_position(
            run_id=f"{run_id}-close",
            pair="BTC/USDC",
            timeframe="4h",
            position_id=position.id,
            exit_price=69000.0,
            reason="take_profit_hit",
        )
        self.repo.refresh_performance(run_id=f"{run_id}-close", pair="BTC/USDC", timeframe="4h")
        self.repo.record_notification(
            run_id=run_id,
            notification_type="PIPELINE_FINISH",
            sent_ok=False,
            error_message="telegram_disabled",
        )

        reports_dir = self.tmp_path / "reports"
        exporter = DashboardExporter(repository=self.repo, reports_dir=reports_dir)
        paths = exporter.export()

        required = [
            "xlsx",
            "pipeline_runs_csv",
            "agent_outputs_csv",
            "trades_csv",
            "active_positions_csv",
            "performance_csv",
            "indicators_csv",
        ]
        for key in required:
            self.assertTrue(paths[key].exists(), f"missing output for {key}")
            self.assertGreater(paths[key].stat().st_size, 0, f"empty output for {key}")

        workbook = load_workbook(paths["xlsx"])
        self.assertEqual(
            workbook.sheetnames,
            ["Runs", "Agents", "Indicators", "Trades", "ActivePositions", "Performance", "ErrorsWarnings"],
        )

        pipeline_csv_header = paths["pipeline_runs_csv"].read_text(encoding="utf-8").splitlines()[0]
        self.assertIn("run_id", pipeline_csv_header)
        self.assertIn("status", pipeline_csv_header)
        performance_csv_header = paths["performance_csv"].read_text(encoding="utf-8").splitlines()[0]
        self.assertIn("total_trades", performance_csv_header)


if __name__ == "__main__":
    unittest.main()
